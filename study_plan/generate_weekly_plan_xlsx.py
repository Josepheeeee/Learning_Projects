from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).resolve().parent
output_path = BASE_DIR / "2026_weekly_study_checklist.xlsx"

stage_info = [
    ["阶段1", "基础重建", "第1-8周", "2026.03.10 ~ 2026.05.04", "2026.05.04", "B站左神/马士兵算法数据结构全家桶（每周5节） + 3Blue1Brown 线代 + MIT 18.06 前半段 + StatQuest 概率基础", "补齐 DSA、线代、微积分、概率底子，建立刷题与笔记系统"],
    ["阶段2", "ML/DL 主线", "第9-14周", "2026.05.05 ~ 2026.06.15", "2026.06.15", "B站算法课继续每周5节 + PyTorch 官方教程 + D2L + CS231n 前半部分", "算法主线继续推进，同时逐步接入 PyTorch 与视觉基础"],
    ["阶段3", "RL 与动力学", "第15-20周", "2026.06.16 ~ 2026.07.27", "2026.07.27", "David Silver RL + Sutton & Barto 前半本 + Gymnasium 文档 + PlaNet/Dreamer 解读", "完成 RL 入门，做出 dynamics model + planning 的项目1初版"],
    ["阶段4", "仿真接入", "第21-26周", "2026.07.28 ~ 2026.09.07", "2026.09.07", "MuJoCo 官方文档 + MIT Underactuated Robotics 前几讲 + 表示学习材料", "完成二维物理 world model 项目第一版，并接入仿真环境"],
    ["阶段5", "前沿轻量切入", "第27-30周", "2026.09.08 ~ 2026.10.05", "2026.10.05", "DreamerV2/V3 解读 + JEPA 讲解材料 + 可选 PINN/JAX 入门", "理解前沿思想，但不陷入过深复现"],
    ["阶段6", "求职准备", "第31-34周", "2026.10.06 ~ 2026.11.02", "2026.11.02", "回看 CS229 / PyTorch / RL 基础 + 简历/README/表达材料", "项目定稿、简历打磨、开始投递"],
    ["阶段7", "冲刺 offer", "第35-38周", "2026.11.03 ~ 2026.11.30", "2026.11.30", "LeetCode 高频题单 + ML/RL 面试问答清单 + 项目模拟面试", "高频复盘、模拟面试、持续投递"],
]

weeks = [
    [1, "阶段1", "2026.03.10-2026.03.16", "2026.05.04", "环境搭建 + Python/复杂度复习", "B站算法课：每周5节，推进基础概念、数组、复杂度", "复习 Python、NumPy、时间复杂度；开始数组/链表", "LC 10-12 题：数组、双指针、链表", "复杂度、递归、基本数学符号", "建立学习仓库与笔记体系", "GitHub 学习仓库初始化 + 第1周复盘", 36],
    [2, "阶段1", "2026.03.17-2026.03.23", "2026.05.04", "栈/队列/哈希表", "B站算法课：每周5节，推进栈、队列、哈希表", "掌握常见模板与哈希思想", "LC 10-12 题：栈、队列、哈希", "离散思维与集合映射", "手写常见数据结构", "数据结构手写笔记 v1", 36],
    [3, "阶段1", "2026.03.24-2026.03.30", "2026.05.04", "树与递归", "B站算法课：每周5节，推进树、递归；3Blue1Brown：向量、矩阵", "二叉树遍历、递归思维", "LC 8-10 题：树遍历、递归", "线代：向量、矩阵、线性变换", "用 Python 实现树遍历模板", "树题总结 + 线代笔记1", 38],
    [4, "阶段1", "2026.03.31-2026.04.06", "2026.05.04", "堆/优先队列/并查集", "B站算法课：每周5节，推进堆、并查集；StatQuest：概率基础", "理解堆的应用场景与并查集", "LC 8-10 题：堆、并查集", "概率基础：条件概率、期望、方差", "实现小顶堆/并查集", "常见结构模板总结", 38],
    [5, "阶段1", "2026.04.07-2026.04.13", "2026.05.04", "图论入门", "B站算法课：每周5节，推进图、BFS、DFS；微积分入门", "掌握 BFS、DFS、visited 模式", "LC 8-10 题：图遍历", "微积分：导数、链式法则、偏导", "实现图的邻接表与遍历", "图论模板总结", 40],
    [6, "阶段1", "2026.04.14-2026.04.20", "2026.05.04", "最短路/拓扑排序", "B站算法课：每周5节，推进最短路、拓扑；MIT 18.06：特征值、SVD 直觉", "理解有向图、拓扑、最短路思想", "LC 8-10 题：拓扑、最短路", "线代：特征值、SVD 建立直觉", "用 NumPy 实现线性回归/逻辑回归", "第一个小项目 README", 40],
    [7, "阶段1", "2026.04.21-2026.04.27", "2026.05.04", "二分/贪心", "B站算法课：每周5节，推进二分、贪心；统计补充：MLE、正则化", "识别单调性与贪心选择", "LC 8-10 题：二分、贪心", "统计直觉：MLE、过拟合、正则化", "整理错题与模板库", "累计 60 题左右", 38],
    [8, "阶段1", "2026.04.28-2026.05.04", "2026.05.04", "回溯/DP 入门", "B站算法课：每周5节，推进回溯、DP 入门", "掌握状态设计与搜索剪枝", "LC 8-10 题：回溯、DP 入门", "阶段数学复盘", "完成阶段总结文档", "阶段1总结 + 技能盘点", 40],
    [9, "阶段2", "2026.05.05-2026.05.11", "2026.06.15", "算法主线继续 + ML 延后接入", "B站算法课：每周5节；PyTorch 轻量入门开始接触", "算法继续推进，同时开始接触建模基础", "LC 6-8 题保持手感", "监督学习框架基础概念", "从零写逻辑回归", "逻辑回归代码 + 笔记", 38],
    [10, "阶段2", "2026.05.12-2026.05.18", "2026.06.15", "算法主线继续 + 轻量 ML 基础", "B站算法课：每周5节；机器学习基础概念补充", "学习评估指标、基础建模思路", "LC 6-8 题", "偏差-方差、交叉验证、正则化", "比较多个传统模型效果", "ML baseline 对比图", 38],
    [11, "阶段2", "2026.05.19-2026.05.25", "2026.06.15", "PyTorch 入门", "PyTorch 官方教程：tensor、autograd、训练循环", "张量、自动求导、训练循环", "LC 6-8 题", "计算图与梯度传播", "用 PyTorch 拟合简单函数", "PyTorch 训练模板", 40],
    [12, "阶段2", "2026.05.26-2026.06.01", "2026.06.15", "深度学习基础", "D2L/PyTorch：MLP、优化器、MNIST", "MLP、优化器、正则化", "LC 6-8 题", "SGD/Adam、batch、过拟合", "完成 MNIST 分类器", "DL 项目 1：MNIST", 40],
    [13, "阶段2", "2026.06.02-2026.06.08", "2026.06.15", "视觉基础", "CS231n：CNN、卷积、池化", "学习 CNN、卷积与池化", "LC 6-8 题", "视觉基础与训练技巧", "跑通 CIFAR-10 基础分类", "CNN baseline", 40],
    [14, "阶段2", "2026.06.09-2026.06.15", "2026.06.15", "序列建模 + world model 入门", "CS231n 训练技巧 + World Models 入门材料", "了解 RNN/LSTM；读 World Models", "LC 6-8 题", "时序建模基础", "完成 DL 基础项目整理", "DL 项目 1 定稿", 38],
    [15, "阶段3", "2026.06.16-2026.06.22", "2026.07.27", "强化学习基础 1", "David Silver RL：MDP、policy、reward", "MDP、state、action、reward、policy", "LC 5-6 题", "RL 基本概念", "安装 Gymnasium 并跑通 CartPole", "RL 入门环境跑通", 38],
    [16, "阶段3", "2026.06.23-2026.06.29", "2026.07.27", "强化学习基础 2", "David Silver RL：value function、Bellman、Q-learning", "价值函数、Bellman、Q-learning 思想", "LC 5-6 题", "value / policy 基础", "写 RL 基础笔记", "RL 基础总结图", 38],
    [17, "阶段3", "2026.06.30-2026.07.06", "2026.07.27", "model-based RL 概念", "PlaNet/Dreamer 解读：dynamics model、planning、rollout", "理解 dynamics model、planning、rollout", "LC 5-6 题", "model-based RL 核心概念", "画出 world model 流程图", "world model 笔记 v1", 40],
    [18, "阶段3", "2026.07.07-2026.07.13", "2026.07.27", "项目1启动：一步预测", "Gymnasium 文档 + 项目1实践", "收集环境数据并训练 s,a->s_next", "LC 5-6 题", "监督学习视角下的 dynamics modeling", "完成项目1数据采样与训练脚本", "项目1 第一版结果", 42],
    [19, "阶段3", "2026.07.14-2026.07.20", "2026.07.27", "项目1：多步 rollout", "Dreamer 解读：rollout 与误差累积", "分析误差累积与 rollout 稳定性", "LC 5-6 题", "多步预测误差分析", "加入多步预测与可视化", "项目1 误差分析图", 42],
    [20, "阶段3", "2026.07.21-2026.07.27", "2026.07.27", "项目1：planning", "RL 补充材料：MPC / random shooting 基础", "加入 random shooting 或简化 MPC", "LC 5-6 题", "planning 基础", "完成控制闭环实验", "项目1 README 初版", 42],
    [21, "阶段4", "2026.07.28-2026.08.03", "2026.09.07", "MuJoCo 入门", "MuJoCo 官方文档：环境、状态、动作、reward", "理解状态、动作、奖励与连续控制", "LC 5-6 题", "MuJoCo 文档入门", "跑通一个简单 MuJoCo 环境", "MuJoCo 环境记录", 40],
    [22, "阶段4", "2026.08.04-2026.08.10", "2026.09.07", "项目1强化", "MuJoCo 示例：连续控制环境实践", "将项目1迁移到更连续的环境", "LC 5-6 题", "model-free 与 model-based 对比", "记录 reward 与预测误差", "项目1 v2", 42],
    [23, "阶段4", "2026.08.11-2026.08.17", "2026.09.07", "项目2启动：二维物理数据", "自建数据集 + 简单物理规则资料", "设计小球碰撞/重力反弹数据集", "LC 5-6 题", "简单物理规则与数据生成", "完成数据生成脚本 + 可视化", "项目2 数据集", 42],
    [24, "阶段4", "2026.08.18-2026.08.24", "2026.09.07", "项目2：状态预测", "时序预测相关资料：未来状态预测 baseline", "先做最简单的状态预测版本", "LC 5-6 题", "预测任务与损失设计", "训练未来轨迹预测模型", "项目2 baseline", 42],
    [25, "阶段4", "2026.08.25-2026.08.31", "2026.09.07", "项目2：latent world model", "表示学习/encoder 教程：latent representation", "加入编码器与 latent dynamics", "LC 5-6 题", "表示学习基础", "完成 latent rollout 可视化", "项目2 latent 版", 42],
    [26, "阶段4", "2026.09.01-2026.09.07", "2026.09.07", "项目2：结果分析", "项目分析资料：rollout 误差与消融分析", "对比像素/状态/latent 预测优劣", "LC 5-6 题", "误差来源与物理一致性", "补实验图、消融和 README", "项目2 第一版定稿", 42],
    [27, "阶段5", "2026.09.08-2026.09.14", "2026.10.05", "Dreamer 系列理解", "DreamerV2/V3 解读：representation、dynamics、imagination", "理解 representation / dynamics / imagination", "LC 5-6 题", "前沿架构理解", "整理 Dreamer 架构图", "前沿笔记 1", 38],
    [28, "阶段5", "2026.09.15-2026.09.21", "2026.10.05", "JEPA 核心思想", "JEPA 演讲/博客/论文摘要：latent prediction", "理解 latent prediction vs autoregressive", "LC 5-6 题", "JEPA 基本思想", "写 JEPA 对比笔记", "前沿笔记 2", 38],
    [29, "阶段5", "2026.09.22-2026.09.28", "2026.10.05", "二选一：PINN 或继续打磨项目2", "PINN 入门或 JAX 入门；否则继续项目优化", "若进度顺利做 1D PINN，否则继续强化 world model 项目", "LC 5-6 题", "PDE / physics loss（可选）", "完成加分项或项目优化", "可选加分成果", 40],
    [30, "阶段5", "2026.09.29-2026.10.05", "2026.10.05", "简历准备", "简历/GitHub README 教程：项目整理与表达", "把两个主项目整理成可讲版本", "LC 5-6 题", "项目叙事与表达", "准备中英文简历初稿", "简历 v1", 38],
    [31, "阶段6", "2026.10.06-2026.10.12", "2026.11.02", "项目1定稿", "面试知识点回看：CS229 + PyTorch", "完善 README、图表、复现步骤", "LC 6-8 题", "项目讲述逻辑", "补单元测试、补结果说明", "项目1 最终版", 40],
    [32, "阶段6", "2026.10.13-2026.10.19", "2026.11.02", "项目2定稿", "面试知识点回看：RL + world model 项目回顾", "完善结果展示与失败分析", "LC 6-8 题", "实验设计与 ablation", "完成最终项目主页", "项目2 最终版", 40],
    [33, "阶段6", "2026.10.20-2026.10.26", "2026.11.02", "开始投递", "JD 反向补课：按岗位需求查漏补缺", "整理岗位表并批量投递", "LC 6-8 题", "岗位 JD 分析", "建立投递跟踪表", "投递清单 v1", 38],
    [34, "阶段6", "2026.10.27-2026.11.02", "2026.11.02", "模拟面试", "模拟面试材料：项目表达与行为面", "针对项目讲解、算法题、ML 问答演练", "LC 6-8 题", "行为面 + 技术面表达", "录音/录像复盘", "面试问答清单", 40],
    [35, "阶段7", "2026.11.03-2026.11.09", "2026.11.30", "高频算法回炉", "LeetCode 高频题单：图、堆、DP、二分回炉", "集中复习堆、图、二分、DP", "LC 8-10 题：高频中等", "模板与复杂度表达", "整理算法题模板", "算法模板文档", 40],
    [36, "阶段7", "2026.11.10-2026.11.16", "2026.11.30", "ML / DL / RL 问答", "ML / RL 问答清单：高频八股与项目问答", "准备常见问答：过拟合、优化器、world model", "LC 6-8 题", "机器学习八股 + 项目问答", "整理 50 问清单", "ML/RL QA 文档", 38],
    [37, "阶段7", "2026.11.17-2026.11.23", "2026.11.30", "专项准备：世界模型/强化学习", "专项问答笔记：world model / dynamics / planning", "针对 dynamics error、latent model、planning 做深挖", "LC 6-8 题", "专项表达训练", "模拟 15 分钟项目面试", "专项问答定稿", 38],
    [38, "阶段7", "2026.11.24-2026.11.30", "2026.11.30", "终版投递与查漏补缺", "无新课：只做面试、投递、复盘", "集中投递、补漏洞、调整项目叙事", "LC 4-6 题维持手感", "最终复盘", "输出终版简历、终版项目导航页", "求职材料最终版", 36],
]

resources = [
    ["数据结构与算法", "B站左神/马士兵算法数据结构全家桶（每周5节）", "当前唯一主课"],
    ["算法补充", "MIT 6.006 / CS61B / Princeton Algorithms", "查漏补缺，不再作为主线"],
    ["机器学习", "CS229 / 6.036（后置）", "不再与当前算法主课并行硬推"],
    ["深度学习", "CS231n / PyTorch 官方教程 / D2L", "主线必学"],
    ["数学", "MIT 18.06 / 3Blue1Brown / StatQuest", "线代、概率、微积分补底"],
    ["强化学习", "David Silver RL / Sutton & Barto / Spinning Up", "主线必学"],
    ["仿真与控制", "MuJoCo Docs / MIT Underactuated Robotics", "中后期"],
    ["世界模型", "World Models / PlaNet / Dreamer 系列", "中后期主线"],
    ["前沿阅读", "JEPA 相关文章", "只做轻量理解"],
    ["加分项", "PINN / JAX", "有余力再做"],
]

worldmodel_projects = [
    ["P1", "StateWorld-Sim", "二维物理状态模拟器", "低-中", "1-2 周", "自己实现 2D 小世界、碰撞、重力、rollout", "world model 地基，训练状态表示与规则建模"],
    ["P2", "StateWorld-Predictor", "下一状态预测器", "中", "1-2 周", "用模拟器数据做 s_t -> s_t+1 预测与多步 rollout", "第一个真正的 world model 味道项目"],
    ["P3", "State2Frame-Renderer", "状态到图像序列渲染器", "中", "1 周", "把状态序列渲染成帧序列，形成可控视频数据", "从状态世界过渡到图像/视频世界"],
    ["P4", "FrameWorld-Baseline", "图像未来帧预测基线", "中-高", "2-3 周", "输入前几帧预测下一帧或未来帧", "正式进入视频型 world model"],
    ["P5", "Mini-Latent-WorldModel", "轻量 latent world model", "高", "3-4 周", "编码器 + latent transition + 解码器", "最关键的主项目，连接现代 world model 思维"],
    ["P6", "Video-WorldModel-Playground", "轻量视频世界模型实验场", "高", "3-6 周", "参考 Open-Sora 思想做最小视频预测/生成实验", "后期升级项目，不建议立即做"],
]

github_watchlist = [
    ["CompVis/stable-diffusion", "重点读", "latent diffusion / encoder-decoder / UNet / conditioning", "热度高，适合理解 latent generation，不作为当前主项目"],
    ["hpcaitech/Open-Sora", "重点读", "video generation / 3D-VAE / 时空建模 / diffusion transformer", "视频世界模型视野核心仓库，先读后轻量实验"],
    ["danijar/dreamerv3", "重点读", "control-style world model / imagined trajectories / actor-critic", "主线思想仓库，帮助你理解正统 world model"],
    ["MagicDrive（关键词跟踪）", "后续补充", "autonomous driving world model / 条件视频生成", "先做关键词调研，不作为近期主项目"],
    ["Emu（关键词跟踪）", "后续补充", "多模态生成 / vision-language foundation model", "先做观察对象，不作为当前主线"],
]

next_8_weeks = [
    ["第1-2周", "StateWorld-Sim", "搭建二维状态世界、碰撞与重力、生成 rollout", "读 DreamerV3 README 与方法图", "1页笔记：world model 的核心模块"],
    ["第3-4周", "StateWorld-Predictor", "做一步预测、多步 rollout、误差分析", "读 Stable Diffusion README 与 latent diffusion 基本结构", "1页笔记：什么是 latent space generation"],
    ["第5周", "State2Frame-Renderer", "把状态序列渲染成帧序列", "读 Open-Sora README/报告概览", "输出时空建模流程图"],
    ["第6-7周", "FrameWorld-Baseline", "做图像未来帧预测 baseline", "继续 Open-Sora 配置结构 + Dreamer imagined rollout", "输出图像 world model baseline 复盘"],
    ["第8周", "Mini-Latent-WorldModel 启动", "开始做 encoder + latent transition 的最小版本", "MagicDrive / Emu 关键词调研", "记录驾驶 world model 与多模态生成的区别"],
]

first_8_weeks_detail = [
    [1, "当前进度后连续5节", "数组、复杂度、二分基础、递归入门、简单排序思想", "两数之和；合并两个有序链表；有效的括号；二叉树的中序遍历；买卖股票的最佳时机；最大子数组和", "数组前半：二分查找、移除元素、有序数组的平方、长度最小的子数组", "路径搜索可视化：建目录、画网格、定义起终点和障碍、跑通 BFS", "最小可运行网格搜索脚本 + BFS 笔记"],
    [2, "再连续5节", "链表、栈、队列、哈希表、双指针", "反转链表；环形链表；最长连续序列；三数之和；盛最多水的容器；无重复字符的最长子串", "链表前半：移除链表元素、反转链表、两两交换节点、删除倒数第N个结点", "路径搜索可视化：加入 DFS、随机障碍、地图编辑、对比 BFS/DFS", "BFS/DFS 对比可视化 + README 初稿"],
    [3, "再连续5节", "树、堆、优先队列、递归深入、二叉树遍历", "二叉树的层序遍历；验证二叉搜索树；对称二叉树；翻转二叉树；二叉树的最大深度；从前序与中序构造二叉树", "二叉树前半：前中后序遍历、层序遍历、翻转二叉树、对称二叉树", "路径搜索项目收尾：加入 Dijkstra；项目2启动：任务调度/拓扑排序系统，定义任务节点格式", "路径搜索项目 v1 + 算法对比图"],
    [4, "再连续5节", "图基础、BFS、DFS、拓扑排序、并查集/连通性", "岛屿数量；课程表；腐烂的橘子；克隆图；实现 Trie；单词搜索", "图论前半：岛屿问题、所有可达路径、并查集入门、拓扑排序相关题", "任务调度系统：输入依赖图、输出执行顺序、加入优先级、可视化依赖", "调度系统最小版 + 拓扑排序笔记"],
    [5, "再连续5节", "最短路、贪心、图优化、堆在图中的应用、DP 引入", "零钱兑换；爬楼梯；打家劫舍；最长递增子序列；组合总和；跳跃游戏", "贪心前半：分发饼干、摆动序列、最大子数组和、跳跃游戏", "任务调度系统收尾；项目3启动：二维物理规则模拟器，先定义状态和边界", "调度系统 v1 + 物理模拟器状态设计文档"],
    [6, "再连续5节", "DP 基础、状态定义、转移方程、记忆化/递推、经典DP 入门", "不同路径；最小路径和；编辑距离；最长公共子序列；分割等和子集；单词拆分", "DP 前半：斐波那契、爬楼梯、最小花费爬楼梯、不同路径", "二维物理规则模拟器：小球移动、边界碰撞、时间步更新、多次 rollout 保存", "可连续播放的小球运动模拟器 + rollout 笔记"],
    [7, "再连续5节", "DP 进阶、背包、区间/序列DP、递归到DP的转换、综合题型", "最长回文子串；最长有效括号；乘积最大子数组；完全平方数；目标和；单词拆分（复刷）", "DP 后半：01背包基础、分割等和子集、最后一块石头的重量II、目标和", "二维物理规则模拟器：加入重力、多个物体、简化碰撞、导出状态序列", "可导出数据的物理模拟器 v1"],
    [8, "再连续5节", "综合复盘、高频题型、复杂度复习、模板归纳、阶段总结", "LRU 缓存；搜索二维矩阵II；接雨水；旋转图像；合并区间；每日温度", "回顾前7周错题；图/树/DP 各补1-2题", "项目3收尾；项目4启动：StateWorld-Predictor，确定输入输出并做一步预测 baseline", "阶段1项目总复盘 + world model 项目准备"],
]

wb = Workbook()
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9D9D9")


def style_sheet(ws, header_row=1):
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# Sheet 1: 每周学习清单
ws = wb.active
ws.title = "每周学习清单"
headers = [
    "周次", "阶段", "周日期", "阶段预期完成时间", "学习主题", "本周课程建议", "本周重点",
    "算法任务", "数学/理论任务", "项目任务", "本周输出", "建议投入(小时)", "完成状态", "实际完成度", "备注"
]
ws.append(headers)
for row in weeks:
    ws.append(row + ["未开始", "", ""])
style_sheet(ws)
widths = {
    "A": 8, "B": 10, "C": 18, "D": 16, "E": 24, "F": 34, "G": 28,
    "H": 22, "I": 24, "J": 28, "K": 24, "L": 12, "M": 12, "N": 12, "O": 18
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:O{ws.max_row}"
status_dv = DataValidation(type="list", formula1='"未开始,进行中,已完成,延期"', allow_blank=True)
progress_dv = DataValidation(type="list", formula1='"0%,25%,50%,75%,100%"', allow_blank=True)
ws.add_data_validation(status_dv)
ws.add_data_validation(progress_dv)
status_dv.add(f"M2:M{ws.max_row}")
progress_dv.add(f"N2:N{ws.max_row}")

# Sheet 2: 阶段规划
stage_ws = wb.create_sheet("阶段规划")
stage_headers = ["阶段", "名称", "周次范围", "时间范围", "预期完成时间", "阶段课程建议", "阶段目标"]
stage_ws.append(stage_headers)
for row in stage_info:
    stage_ws.append(row)
style_sheet(stage_ws)
for col, width in {"A": 10, "B": 14, "C": 12, "D": 24, "E": 16, "F": 58, "G": 36}.items():
    stage_ws.column_dimensions[col].width = width
stage_ws.freeze_panes = "A2"

# Sheet 3: 执行说明
summary = wb.create_sheet("执行说明")
summary_rows = [
    ["项目", "内容"],
    ["总目标", "2026 年 12 月前完成算法/强化学习/世界模型导向的求职准备"],
    ["默认起始时间", "2026-03-10；若实际开始更晚，可整体顺延，但不建议压缩阶段内容"],
    ["当前算法主课", "B站左神/马士兵算法数据结构全家桶；每周固定学5节"],
    ["推荐日均学习时长", "5-7 小时"],
    ["每日建议结构", "算法 1.5-2h + 课程 2h + 项目 1.5-2h + 复盘 0.5h"],
    ["核心成果", "2 个主项目 + 150 题左右算法题 + 一套可投递简历"],
    ["主项目 1", "经典控制环境中的 dynamics model + MPC"],
    ["主项目 2", "二维物理世界的小型 world model"],
    ["加分项", "MuJoCo 简化 model-based RL / 1D PINN"],
    ["使用方式", "每周开始前看‘本周课程建议’和‘本周重点’；每周结束更新完成状态、完成度、备注"],
]
for row in summary_rows:
    summary.append(row)
style_sheet(summary)
summary.column_dimensions["A"].width = 18
summary.column_dimensions["B"].width = 92

# Sheet 4: 资源清单
res = wb.create_sheet("资源清单")
res.append(["类别", "资源", "备注"])
for row in resources:
    res.append(row)
style_sheet(res)
res.column_dimensions["A"].width = 18
res.column_dimensions["B"].width = 52
res.column_dimensions["C"].width = 24

# Sheet 5: WorldModel项目路线
proj = wb.create_sheet("WorldModel项目路线")
proj.append(["编号", "项目代号", "项目名称", "难度", "预计时间", "核心任务", "定位"])
for row in worldmodel_projects:
    proj.append(row)
style_sheet(proj)
for col, width in {"A": 8, "B": 22, "C": 24, "D": 10, "E": 12, "F": 44, "G": 40}.items():
    proj.column_dimensions[col].width = width
proj.freeze_panes = "A2"

# Sheet 6: GitHub仓库观察
gh = wb.create_sheet("GitHub仓库观察")
gh.append(["仓库/方向", "当前建议", "你该重点看什么", "备注"])
for row in github_watchlist:
    gh.append(row)
style_sheet(gh)
for col, width in {"A": 28, "B": 14, "C": 44, "D": 40}.items():
    gh.column_dimensions[col].width = width
gh.freeze_panes = "A2"

# Sheet 7: 接下来8周专项
plan8 = wb.create_sheet("接下来8周专项")
plan8.append(["阶段", "主线项目", "主线任务", "副线仓库阅读", "阶段输出"])
for row in next_8_weeks:
    plan8.append(row)
style_sheet(plan8)
for col, width in {"A": 12, "B": 28, "C": 40, "D": 40, "E": 36}.items():
    plan8.column_dimensions[col].width = width
plan8.freeze_panes = "A2"

# Sheet 8: 前8周细化
detail8 = wb.create_sheet("前8周细化")
detail8.append(["周次", "B站课进度", "本周核心主题", "Hot100", "代码随想录", "项目推进", "本周输出"])
for row in first_8_weeks_detail:
    detail8.append(row)
style_sheet(detail8)
for col, width in {"A": 8, "B": 16, "C": 28, "D": 52, "E": 42, "F": 42, "G": 30}.items():
    detail8.column_dimensions[col].width = width
detail8.freeze_panes = "A2"

wb.save(output_path)
print(output_path)
